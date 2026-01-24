import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'PersonaBuddy.settings')
django.setup()
from agent.models import *
import random
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def process(pid):
    records = Record.objects.filter(pid=pid, is_filter=True)
    records_filter = records.filter(filter_result=True)
    records_not_filter = records.filter(filter_result=False)
    print(f"{pid}:", len(records_filter), len(records_not_filter))

    # 随机选10个过滤内容
    records_filter_sample = records_filter.order_by('?')[:10]
    records_not_filter_sample = records_not_filter.order_by('?')[:10]

    return records_filter_sample, records_not_filter_sample

def add_data_validation_l(file_path, sheet_name, column_letter, options):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', showDropDown=True)
    
    for row in range(2, ws.max_row + 1):
        cell = f"{column_letter}{row}"
        dv.add(ws[cell])
    
    ws.add_data_validation(dv)
    
    print(f"Data validation added to {file_path} in column {column_letter}")
    wb.save(file_path)

if __name__ == "__main__":
    all_users = UserPid.objects.all()
    for user in all_users:
        pid = user.pid
        records_filter_sample, records_not_filter_sample = process(pid)
        note_data = []
        for record in records_filter_sample:
            note_data.append([record.title, record.context, record.filter_reason, record.content, ""])
        note_data_df = pd.DataFrame(note_data, columns=["问题标题", "命中规则", "过滤解释", "相关内容", "是否符合您的意愿"])
        note_data_df.to_csv(os.path.join("logs", f"{pid}_filter.csv"), index=False) 
        note_data_df.to_excel(os.path.join("logs", f"{pid}_filter.xlsx"), index=False)
        add_data_validation_l(os.path.join("logs", f"{pid}_filter.xlsx"), "Sheet1", "E", ["是", "否"])

        note_data = []
        for record in records_not_filter_sample:
            note_data.append([record.title, record.filter_reason, record.content, ""])  
        note_data_df = pd.DataFrame(note_data, columns=["问题标题", "不过滤解释", "相关内容", "是否符合您的意愿"])
        note_data_df.to_csv(os.path.join("logs", f"{pid}_not_filter.csv"), index=False) 
        note_data_df.to_excel(os.path.join("logs", f"{pid}_not_filter.xlsx"), index=False)
        add_data_validation_l(os.path.join("logs", f"{pid}_not_filter.xlsx"), "Sheet1", "D", ["是", "否"])

